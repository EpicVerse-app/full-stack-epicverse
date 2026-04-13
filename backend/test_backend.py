"""
EpicVerse Backend — Comprehensive Combo Test Suite (Excel-based)
Tests 50+ combos across different card numbers, modes, and question languages.
Characters: 1-24  |  Attributes: 25-104  |  Modes: 1-7
"""

import openpyxl
from pathlib import Path

DATA_DIR = Path("E:/kriyora/EpicVerse/backend/data")

# ── Q1 & Q2 templates in 10 languages ───────────────────────────────────────
Q1 = {
    "English":    "Is combo {c1}+{c2} valid?",
    "Tamil":      "{c1}+{c2} சேர்க்கை செல்லுபடியாகுமா?",
    "Telugu":     "{c1}+{c2} కాంబో చెల్లుబాటు అవుతుందా?",
    "Hindi":      "क्या {c1}+{c2} कॉम्बो मान्य है?",
    "Spanish":    "¿Es válida la combo {c1}+{c2}?",
    "Japanese":   "{c1}+{c2}のコンボは有効ですか？",
    "French":     "La combo {c1}+{c2} est-elle valide?",
    "German":     "Ist Kombination {c1}+{c2} gültig?",
    "Arabic":     "هل مزيج {c1}+{c2} صالح؟",
    "Portuguese": "A combo {c1}+{c2} é válida?",
}

Q2 = {
    "English":    "Why is this combo valid/invalid?",
    "Tamil":      "இது ஏன் செல்லுபடியாகிறது/இல்லை?",
    "Telugu":     "ఇది ఎందుకు చెల్లుబాటు/కాదు?",
    "Hindi":      "यह कॉम्बो मान्य/अमान्य क्यों है?",
    "Spanish":    "¿Por qué esta combo es válida/inválida?",
    "Japanese":   "このコンボはなぜ有効/無効ですか？",
    "French":     "Pourquoi cette combo est-elle valide/invalide?",
    "German":     "Warum ist diese Kombination gültig/ungültig?",
    "Arabic":     "لماذا هذا المزيج صالح/غير صالح؟",
    "Portuguese": "Por que esta combo é válida/inválida?",
}

LANGS = list(Q1.keys())

# ── Load all mode Excel files into a lookup dict ─────────────────────────────
# key: (char_card_no, attr_card_no, mode_number) → {status, message, char, attr, mode_name}
DB: dict[tuple, dict] = {}
MODE_NAMES: dict[int, str] = {}

for mode_num in range(1, 8):
    path = DATA_DIR / f"mode{mode_num}.xlsx"
    if not path.exists():
        continue
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def ci(name):
        return hdrs.index(name) + 1

    char_ci   = ci("Character Card No.")
    attr_ci   = ci("Attribute Card No.")
    seg_ci    = ci("Final Segment")
    msg_ci    = ci("App Message (Crisp)")
    char_n_ci = ci("Character")
    attr_n_ci = ci("Attribute")
    mode_n_ci = ci("Game Play Mode")

    for row in ws.iter_rows(min_row=2, values_only=True):
        # row is a tuple of values (0-indexed)
        char_no  = row[char_ci - 1]
        attr_no  = row[attr_ci - 1]
        status   = row[seg_ci - 1]
        message  = row[msg_ci - 1]
        char_nm  = row[char_n_ci - 1]
        attr_nm  = row[attr_n_ci - 1]
        mode_nm  = row[mode_n_ci - 1]

        if char_no is None or attr_no is None:
            continue
        key = (int(char_no), int(attr_no), mode_num)
        DB[key]  = {
            "status":    status or "Unknown",
            "message":   message or "",
            "character": char_nm or "—",
            "attribute": attr_nm or "—",
            "mode_name": mode_nm or f"Mode {mode_num}",
        }
        if mode_num not in MODE_NAMES and mode_nm:
            MODE_NAMES[mode_num] = str(mode_nm)

    wb.close()

print(f"Loaded {len(DB)} combo records across {len(MODE_NAMES)} modes.\n")
print("Mode names:")
for k, v in sorted(MODE_NAMES.items()):
    print(f"  Mode {k}: {v}")
print()

# ── 50 test cases: (char_card, attr_card, mode_num, language) ────────────────
# Covering valid, invalid, excluded combos; all 7 modes; all 10 languages
TEST_CASES = [
    # --- 1+29 combo in all 10 languages (Mode 1) ---
    ( 1, 29, 1, "English"),
    ( 1, 29, 1, "Tamil"),
    ( 1, 29, 1, "Hindi"),
    ( 1, 29, 1, "Spanish"),
    ( 1, 29, 1, "Japanese"),
    ( 1, 29, 1, "French"),
    ( 1, 29, 1, "Telugu"),
    ( 1, 29, 1, "German"),
    ( 1, 29, 1, "Arabic"),
    ( 1, 29, 1, "Portuguese"),
    # --- Same combo across modes ---
    ( 1, 29, 2, "English"),
    ( 1, 29, 3, "Tamil"),
    ( 1, 29, 4, "Hindi"),
    ( 1, 29, 5, "Spanish"),
    ( 1, 29, 6, "Japanese"),
    ( 1, 29, 7, "French"),
    # --- Different combos, Mode 1, varied languages ---
    ( 1, 30, 1, "English"),
    ( 2, 29, 1, "Tamil"),
    ( 3, 45, 1, "Hindi"),
    ( 4, 50, 2, "Spanish"),
    ( 5, 33, 3, "Japanese"),
    ( 6, 40, 4, "French"),
    ( 7, 55, 1, "Telugu"),
    ( 8, 60, 2, "English"),
    ( 9, 70, 3, "German"),
    (10, 80, 4, "Arabic"),
    (11, 25, 5, "English"),
    (12, 35, 6, "Tamil"),
    (13, 48, 7, "Hindi"),
    (14, 52, 1, "Spanish"),
    (15, 65, 2, "Japanese"),
    (16, 38, 3, "French"),
    (17, 42, 4, "Telugu"),
    (18, 58, 5, "English"),
    (19, 72, 6, "German"),
    (20, 85, 7, "Arabic"),
    (21, 31, 1, "English"),
    (22, 44, 2, "Tamil"),
    (23, 56, 3, "Hindi"),
    (24, 90, 4, "Spanish"),
    # --- Wider attribute range ---
    ( 1,100, 1, "English"),
    ( 2,104, 2, "Tamil"),
    ( 5, 95, 3, "Spanish"),
    ( 7, 88, 4, "French"),
    ( 1, 75, 5, "English"),
    ( 3, 60, 6, "Telugu"),
    ( 6, 77, 7, "Japanese"),
    ( 9, 66, 1, "German"),
    (12, 55, 2, "Arabic"),
    (15, 44, 3, "Portuguese"),
]


def trunc(text: str, n: int = 55) -> str:
    if not text:
        return "—"
    text = str(text).replace("\n", " ").strip()
    return text[:n] + "…" if len(text) > n else text


# ── Build result rows ─────────────────────────────────────────────────────────
ROWS = []
row_no = 0

for idx, (c1, c2, mode, lang) in enumerate(TEST_CASES):
    rec  = DB.get((c1, c2, mode))
    status   = rec["status"]    if rec else "Not in Excel"
    char_nm  = rec["character"] if rec else "—"
    attr_nm  = rec["attribute"] if rec else "—"
    message  = trunc(rec["message"] if rec else "")
    mode_tag = f"Mode {mode}"

    q1_text = Q1[lang].format(c1=c1, c2=c2)
    q2_text = Q2[lang]

    row_no += 1
    ROWS.append({
        "#":          row_no,
        "Q":          "Q1",
        "Lang":       lang,
        "Mode":       mode_tag,
        "Cards":      f"{c1}+{c2}",
        "Character":  char_nm,
        "Attribute":  attr_nm,
        "Status":     status,
        "Question":   trunc(q1_text, 50),
        "DB Answer":  message,
    })

    row_no += 1
    ROWS.append({
        "#":          row_no,
        "Q":          "Q2",
        "Lang":       lang,
        "Mode":       mode_tag,
        "Cards":      f"{c1}+{c2}",
        "Character":  char_nm,
        "Attribute":  attr_nm,
        "Status":     status,
        "Question":   trunc(q2_text, 50),
        "DB Answer":  message,   # AI would use this to answer "why"
    })


# ── Print table ───────────────────────────────────────────────────────────────
COLS = ["#", "Q", "Lang", "Mode", "Cards", "Character", "Attribute", "Status", "Question", "DB Answer"]
widths = {c: max(len(c), max(len(str(r[c])) for r in ROWS)) for c in COLS}

sep    = "+-" + "-+-".join("-" * widths[c] for c in COLS) + "-+"
header = "| " + " | ".join(c.ljust(widths[c]) for c in COLS) + " |"

print(sep)
print(header)
print(sep)
for r in ROWS:
    line = "| " + " | ".join(str(r[c]).ljust(widths[c]) for c in COLS) + " |"
    print(line)
print(sep)

# ── Summary ───────────────────────────────────────────────────────────────────
q1_rows = [r for r in ROWS if r["Q"] == "Q1"]
statuses = [r["Status"] for r in q1_rows]

from collections import Counter
counts = Counter(statuses)

print(f"\n{'='*60}")
print(f"SUMMARY — {len(q1_rows)} Q1 + {len(q1_rows)} Q2 = {len(ROWS)} total rows")
print(f"{'='*60}")
for stat, cnt in sorted(counts.items()):
    print(f"  {stat:<22}: {cnt}")
print()

modes_tested  = sorted(set(r["Mode"]  for r in q1_rows))
langs_tested  = sorted(set(r["Lang"]  for r in q1_rows))
cards_tested  = sorted(set(r["Cards"] for r in q1_rows))

print(f"Modes tested  ({len(modes_tested)}): {', '.join(modes_tested)}")
print(f"Languages     ({len(langs_tested)}): {', '.join(langs_tested)}")
print(f"Card combos   ({len(cards_tested)}): {', '.join(cards_tested)}")
