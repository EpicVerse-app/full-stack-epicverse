import openpyxl
from pathlib import Path

# Use local data for testing
DATA_DIR = Path("E:/kriyora/EpicVerse/backend/data")

def test_single_combo(char_no, attr_no, mode_num):
    path = DATA_DIR / f"mode{mode_num}.xlsx"
    if not path.exists():
        print(f"❌ File {path} not found")
        return
    
    print(f"Loading {path}...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Headers: {hdrs}")
    
    char_ci = hdrs.index("Character Card No.") + 1
    attr_ci = hdrs.index("Attribute Card No.") + 1
    msg_ci = hdrs.index("App Message (Crisp)") + 1
    status_ci = hdrs.index("Final Segment") + 1
    
    count = 0
    for r in range(2, ws.max_row + 1):
        c_val = ws.cell(r, char_ci).value
        a_val = ws.cell(r, attr_ci).value
        
        if c_val == char_no and a_val == attr_no:
            status = ws.cell(r, status_ci).value
            message = ws.cell(r, msg_ci).value
            print(f"✅ Found Combo {char_no}+{attr_no}")
            print(f"Status: {status}")
            print(f"Message: {message}")
            return
        count += 1
    
    print(f"❌ Combo {char_no}+{attr_no} not found in {count} rows")
    wb.close()

if __name__ == "__main__":
    # Test Combo 1+29 in Mode 1
    test_single_combo(1, 29, 1)
