"""EpicVerse Quick Test - reads Excel and outputs table"""
import sys
import openpyxl
from pathlib import Path
from collections import Counter

DATA_DIR = Path("E:/kriyora/EpicVerse/backend/data")

Q1 = {
    "English":    "Is combo {c1}+{c2} valid?",
    "Tamil":      "{c1}+{c2} சேர்க்கை செல்லுபடியாகுமா?",
    "Hindi":      "क्या {c1}+{c2} कॉम्बो मान्य है?",
    "Spanish":    "¿Es válida la combo {c1}+{c2}?",
    "Japanese":   "{c1}+{c2}のコンボは有効ですか？",
    "French":     "La combo {c1}+{c2} est-elle valide?",
    "Telugu":     "{c1}+{c2} కాంబో చెల్లుబాటు అవుతుందా?",
    "German":     "Ist Kombination {c1}+{c2} gültig?",
    "Arabic":     "هل مزيج {c1}+{c2} صالح؟",
    "Portuguese": "A combo {c1}+{c2} é válida?",
}
Q2 = {
    "English":    "Why is this combo valid/invalid?",
    "Tamil":      "இது ஏன் செல்லுபடியாகிறது/இல்லை?",
    "Hindi":      "यह कॉम्बो मान्य/अमान्य क्यों है?",
    "Spanish":    "¿Por qué esta combo es válida/inválida?",
    "Japanese":   "このコンボはなぜ有効/無効ですか？",
    "French":     "Pourquoi cette combo est-elle valide/invalide?",
    "Telugu":     "ఇది ఎందుకు చెల్లుబాటు/కాదు?",
    "German":     "Warum ist diese Kombination gültig/ungültig?",
    "Arabic":     "لماذا هذا المزيج صالح/غير صالح؟",
    "Portuguese": "Por que esta combo é válida/inválida?",
}

# Load only 3 modes to stay fast
DB = {}
MODE_NAMES = {}
for mode_num in [1, 2, 3]:
    path = DATA_DIR / f"mode{mode_num}.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    def ci(name):
        return hdrs.index(name) + 1
    for r in range(2, ws.max_row + 1):
        c_no = ws.cell(r, ci("Character Card No.")).value
        a_no = ws.cell(r, ci("Attribute Card No.")).value
        stat = ws.cell(r, ci("Final Segment")).value
        msg  = ws.cell(r, ci("App Message (Crisp)")).value
        c_nm = ws.cell(r, ci("Character")).value
        a_nm = ws.cell(r, ci("Attribute")).value
        m_nm = ws.cell(r, ci("Game Play Mode")).value
        if c_no is None: continue
        DB[(int(c_no), int(a_no), mode_num)] = {
            "status": stat or "Unknown",
            "message": (msg or "")[:55] + ("…" if msg and len(msg) > 55 else ""),
            "character": c_nm or "—",
            "attribute": a_nm or "—",
            "mode_name": m_nm or f"Mode{mode_num}",
        }
        if mode_num not in MODE_NAMES and m_nm:
            MODE_NAMES[mode_num] = str(m_nm)
    wb.close()

sys.stdout.write(f"Loaded {len(DB)} records across 3 modes.\n\n")

# Test cases: (char, attr, mode, lang)
CASES = [
    (1,29,1,"English"),   (1,29,1,"Tamil"),     (1,29,1,"Hindi"),
    (1,29,1,"Spanish"),   (1,29,1,"Japanese"),  (1,29,1,"French"),
    (1,29,1,"Telugu"),    (1,29,1,"German"),    (1,29,1,"Arabic"),
    (1,29,1,"Portuguese"),
    (1,29,2,"English"),   (1,29,3,"Hindi"),
    (1,30,1,"English"),   (2,29,1,"Tamil"),     (3,45,1,"Hindi"),
    (4,50,2,"Spanish"),   (5,33,3,"Japanese"),  (6,40,1,"French"),
    (7,55,1,"Telugu"),    (8,60,2,"English"),   (9,70,3,"German"),
    (10,80,1,"Arabic"),   (11,25,2,"English"),  (12,35,3,"Tamil"),
    (13,48,1,"Hindi"),    (14,52,2,"Spanish"),  (15,65,3,"Japanese"),
    (16,38,1,"French"),   (17,42,2,"Telugu"),   (18,58,3,"English"),
    (19,72,1,"German"),   (20,85,2,"Arabic"),   (21,31,3,"English"),
    (22,44,1,"Tamil"),    (23,56,2,"Hindi"),    (24,90,3,"Spanish"),
    (1,100,1,"English"),  (2,104,2,"Tamil"),    (5,95,3,"Spanish"),
    (7,88,1,"French"),    (1,75,2,"English"),   (3,60,3,"Telugu"),
    (6,77,1,"Japanese"),  (9,66,2,"German"),    (12,55,3,"Arabic"),
    (15,44,1,"Portuguese"),(18,66,2,"English"),  (20,55,3,"Tamil"),
    (22,44,1,"Hindi"),    (24,33,2,"Spanish"),
]

ROWS = []
n = 0
for c1,c2,mode,lang in CASES:
    rec = DB.get((c1,c2,mode))
    stat = rec["status"] if rec else "Not Found"
    char = rec["character"] if rec else "—"
    attr = rec["attribute"] if rec else "—"
    msg  = rec["message"] if rec else "—"
    n+=1; ROWS.append({"#":n,"Q":"Q1","Lang":lang,"Mode":f"M{mode}","Cards":f"{c1}+{c2}","Char":char,"Attr":attr,"Status":stat,"Question":Q1[lang].format(c1=c1,c2=c2)[:50],"Answer":msg})
    n+=1; ROWS.append({"#":n,"Q":"Q2","Lang":lang,"Mode":f"M{mode}","Cards":f"{c1}+{c2}","Char":char,"Attr":attr,"Status":stat,"Question":Q2[lang][:50],"Answer":msg})

COLS = ["#","Q","Lang","Mode","Cards","Char","Attr","Status","Question","Answer"]
W = {c: max(len(c), max(len(str(r[c])) for r in ROWS)) for c in COLS}

sep = "+-"+ "-+-".join("-"*W[c] for c in COLS)+"-+"
hdr = "| "+" | ".join(c.ljust(W[c]) for c in COLS)+" |"
sys.stdout.write(sep+"\n")
sys.stdout.write(hdr+"\n")
sys.stdout.write(sep+"\n")
for r in ROWS:
    sys.stdout.write("| "+" | ".join(str(r[c]).ljust(W[c]) for c in COLS)+" |\n")
sys.stdout.write(sep+"\n")

q1s = [r["Status"] for r in ROWS if r["Q"]=="Q1"]
cnt = Counter(q1s)
sys.stdout.write(f"\nSUMMARY — {len(q1s)} Q1 + {len(q1s)} Q2 = {len(ROWS)} total rows\n")
for s,c in sorted(cnt.items()): sys.stdout.write(f"  {s:<22}: {c}\n")
modes  = sorted(set(r["Mode"] for r in ROWS if r["Q"]=="Q1"))
langs  = sorted(set(r["Lang"] for r in ROWS if r["Q"]=="Q1"))
combos = sorted(set(r["Cards"] for r in ROWS if r["Q"]=="Q1"))
sys.stdout.write(f"\nModes : {', '.join(modes)}\n")
sys.stdout.write(f"Langs : {', '.join(langs)}\n")
sys.stdout.write(f"Combos: {', '.join(combos)}\n")
sys.stdout.flush()
